import pickle

import numpy as np
import openpyxl
import pandas as pd

filePath = ['Cell 1', 'Cell 2']
XJ_battery = {}
count = 0
for i in filePath:
    count += 1
    discharge_voltage, discharge_current, discharge_capacity = [], [], []
    charge_voltage, charge_current, charge_capacity = [], [], []

    discharge_data = i + '/Discharging stage.xlsx'
    charge_data = i + '/Charging stage.xlsx'
    battery = {}
    discharge = openpyxl.load_workbook(discharge_data)
    charge = openpyxl.load_workbook(discharge_data)

    # 获取所有sheet名称
    sheet_names = discharge.sheetnames
    for sheet_name in sheet_names:
        # 选择一个sheet
        d_c = discharge[sheet_name]
        c = charge[sheet_name]
        dv, dc, disc = [], [], []
        cv, cc, capa = [], [], []
        for row in d_c.iter_rows():
            dv.append(row[0].value)
            dc.append(row[1].value)
            disc.append(row[2].value)
        for row in c.iter_rows():
            cv.append(row[0].value)
            cc.append(row[1].value)
            capa.append(row[2].value)
        discharge_voltage.append(np.array(dv))
        discharge_current.append(np.array(dc))
        discharge_capacity.append(np.array(disc))
        charge_voltage.append(np.array(cv))
        charge_current.append(np.array(cc))
        charge_capacity.append(np.array(capa))
    battery = {
        'discharge_voltage': discharge_voltage,
        'discharge_current': discharge_current,
        'discharge_capacity': discharge_capacity,
        'discharge_temperature': [[0]],

        'charge_voltage': charge_voltage,
        'charge_current': charge_current,
        'charge_capacity': charge_capacity,
        'charge_temperature': [[0]],

    }
    XJ_battery['battery_' + str(count)] = battery

with open('../../pkl_data/XJ_battery.pkl', 'wb') as f:
    pickle.dump(XJ_battery, f)

